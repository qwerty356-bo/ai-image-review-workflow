"""
AI图片审核工作流 - Excel/CSV导出
"""

import os
import json
import csv

import config


def load_results():
    """加载审核结果"""
    results_file = os.path.join(config.OUTPUT_DIR, "results.json")
    if not os.path.exists(results_file):
        print(f"结果文件不存在: {results_file}")
        print("请先运行 python batch_run.py 完成审核")
        return []
    with open(results_file, "r", encoding="utf-8") as f:
        return json.load(f)


def export_csv(results):
    """导出CSV"""
    output_path = os.path.join(config.OUTPUT_DIR, "report.csv")
    headers = [
        "诗名", "变体", "文件名", "画风", "画风置信度",
        "质量分", "质量说明", "合规", "合规说明",
        "诗词契合度", "契合说明", "有文字", "文字状态",
        "总分", "等级", "审核状态", "审核时间", "Token消耗", "耗时(ms)", "文件大小(MB)"
    ]

    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for r in results:
            text_status = ""
            if r.get("has_text"):
                tc = r.get("text_correct", "")
                text_status = tc if tc else "有文字"
            else:
                text_status = "无文字"

            writer.writerow([
                r.get("poem_name", ""),
                r.get("variant", ""),
                r.get("file_name", ""),
                r.get("style", ""),
                r.get("style_confidence", 0),
                r.get("quality_score", 0),
                r.get("quality_detail", ""),
                "合规" if r.get("compliant", True) else "不合规",
                r.get("compliance_detail", ""),
                r.get("poem_match_score", 0),
                r.get("poem_match_detail", ""),
                "是" if r.get("has_text") else "否",
                text_status,
                r.get("total_score", 0),
                r.get("grade", ""),
                r.get("review_status", ""),
                r.get("reviewed_at", ""),
                r.get("tokens_used", 0),
                r.get("latency_ms", 0),
                r.get("file_size_mb", 0),
            ])

    print(f"CSV已导出: {output_path}")
    return output_path


def export_excel(results):
    """导出Excel（带格式化）"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("缺少openpyxl库，尝试安装...")
        import subprocess
        subprocess.check_call([
            os.path.join(os.path.dirname(sys.executable), "pip"),
            "install", "openpyxl"
        ])
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

    import sys

    wb = Workbook()

    # ============ Sheet 1: 审核明细 ============
    ws = wb.active
    ws.title = "审核明细"

    headers = [
        "诗名", "变体", "文件名", "画风", "画风置信度",
        "质量分", "质量说明", "合规", "合规说明",
        "诗词契合度", "契合说明", "有文字", "文字状态",
        "总分", "等级", "审核状态", "审核时间", "Token", "耗时(ms)", "大小(MB)"
    ]

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 等级颜色
    grade_fills = {
        "A": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "B": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
        "C": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "D": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }
    grade_fonts = {
        "A": Font(color="006100", bold=True),
        "B": Font(color="0066CC", bold=True),
        "C": Font(color="CC6600", bold=True),
        "D": Font(color="9C0006", bold=True),
    }

    # 数据行
    for row_idx, r in enumerate(results, 2):
        grade = r.get("grade", "")
        text_status = ""
        if r.get("has_text"):
            tc = r.get("text_correct", "")
            text_status = tc if tc else "有文字"
        else:
            text_status = "无文字"

        row_data = [
            r.get("poem_name", ""),
            r.get("variant", ""),
            r.get("file_name", ""),
            r.get("style", ""),
            r.get("style_confidence", 0),
            r.get("quality_score", 0),
            r.get("quality_detail", ""),
            "合规" if r.get("compliant", True) else "不合规",
            r.get("compliance_detail", ""),
            r.get("poem_match_score", 0),
            r.get("poem_match_detail", ""),
            "是" if r.get("has_text") else "否",
            text_status,
            r.get("total_score", 0),
            grade,
            r.get("review_status", ""),
            r.get("reviewed_at", ""),
            r.get("tokens_used", 0),
            r.get("latency_ms", 0),
            r.get("file_size_mb", 0),
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            # 等级列上色
            if col == 15 and grade in grade_fills:  # 等级列
                cell.fill = grade_fills[grade]
                cell.font = grade_fonts[grade]
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # 总分列上色
            elif col == 14 and grade in grade_fills:
                cell.fill = grade_fills[grade]

    # 列宽
    col_widths = [12, 6, 30, 10, 10, 10, 30, 8, 20, 10, 30, 8, 10, 10, 6, 10, 20, 10, 10, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动筛选
    ws.auto_filter.ref = ws.dimensions

    # ============ Sheet 2: 统计概览 ============
    ws2 = wb.create_sheet("统计概览")

    success = [r for r in results if r.get("review_status") == "success"]
    total = len(success)
    grade_counts = {"A": 0, "B": 0, "C": 0, "D": 0}
    style_counts = {}
    for r in success:
        g = r.get("grade", "C")
        grade_counts[g] = grade_counts.get(g, 0) + 1
        style = r.get("style", "未知")
        style_counts[style] = style_counts.get(style, 0) + 1

    avg_quality = sum(r.get("quality_score", 0) for r in success) / total if total else 0
    avg_poem = sum(r.get("poem_match_score", 0) for r in success) / total if total else 0
    avg_total = sum(r.get("total_score", 0) for r in success) / total if total else 0
    total_tokens = sum(r.get("tokens_used", 0) for r in success)
    non_compliant = sum(1 for r in success if not r.get("compliant", True))

    stats_data = [
        ["统计项", "数值"],
        ["图片总数", len(results)],
        ["审核成功", total],
        ["审核失败", len(results) - total],
        ["合规数", total - non_compliant],
        ["不合规数", non_compliant],
        ["合规率", f"{(total - non_compliant) / total * 100:.1f}%" if total else "0%"],
        ["", ""],
        ["A级(≥8.0)", grade_counts.get("A", 0)],
        ["B级(≥6.0)", grade_counts.get("B", 0)],
        ["C级(<6.0)", grade_counts.get("C", 0)],
        ["D级(不合规)", grade_counts.get("D", 0)],
        ["", ""],
        ["平均质量分", round(avg_quality, 2)],
        ["平均契合度", round(avg_poem, 2)],
        ["平均总分", round(avg_total, 2)],
        ["总Token消耗", total_tokens],
    ]

    for row_idx, (label, val) in enumerate(stats_data, 1):
        ws2.cell(row=row_idx, column=1, value=label).font = Font(bold=(row_idx == 1))
        ws2.cell(row=row_idx, column=2, value=val).font = Font(bold=(row_idx == 1))
        if row_idx == 1:
            for col in [1, 2]:
                ws2.cell(row=row_idx, column=col).fill = header_fill
                ws2.cell(row=row_idx, column=col).font = header_font

    # 画风统计
    ws2.cell(row=len(stats_data) + 2, column=1, value="画风分布").font = Font(bold=True, size=14)
    for i, (style, count) in enumerate(sorted(style_counts.items(), key=lambda x: -x[1])):
        r = len(stats_data) + 3 + i
        ws2.cell(row=r, column=1, value=style)
        ws2.cell(row=r, column=2, value=count)
        ws2.cell(row=r, column=3, value=f"{count / total * 100:.1f}%" if total else "")

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 10

    # ============ Sheet 3: 异常图片 ============
    errors = [r for r in results if r.get("review_status") == "error"]
    non_compliant_list = [r for r in success if not r.get("compliant", True)]
    text_issues = [r for r in success if r.get("text_correct") == "有乱码"]

    ws3 = wb.create_sheet("异常清单")

    ws3.cell(row=1, column=1, value="类型").font = header_font
    ws3.cell(row=1, column=2, value="文件名").font = header_font
    ws3.cell(row=1, column=3, value="说明").font = header_font
    for col in [1, 2, 3]:
        ws3.cell(row=1, column=col).fill = header_fill

    row = 2
    for r in errors:
        ws3.cell(row=row, column=1, value="审核失败")
        ws3.cell(row=row, column=2, value=r.get("file_name", ""))
        ws3.cell(row=row, column=3, value=r.get("error_message", ""))
        row += 1

    for r in non_compliant_list:
        ws3.cell(row=row, column=1, value="不合规")
        ws3.cell(row=row, column=2, value=r.get("file_name", ""))
        ws3.cell(row=row, column=3, value=r.get("compliance_detail", ""))
        row += 1

    for r in text_issues:
        ws3.cell(row=row, column=1, value="文字乱码")
        ws3.cell(row=row, column=2, value=r.get("file_name", ""))
        ws3.cell(row=row, column=3, value=f"检测到乱码文字")
        row += 1

    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 35
    ws3.column_dimensions["C"].width = 50

    # 保存
    output_path = os.path.join(config.OUTPUT_DIR, "report.xlsx")
    wb.save(output_path)
    print(f"Excel已导出: {output_path}")
    return output_path


def main():
    """主入口"""
    results = load_results()
    if not results:
        return

    export_csv(results)
    export_excel(results)
    print(f"\n导出完成！文件位于: {config.OUTPUT_DIR}")


if __name__ == "__main__":
    main()
