"""
AI视频审核工作流 - Excel/CSV导出
导出视频聚合结果 + 帧级明细，独立于图片审核导出
"""

import os
import sys
import json
import csv

import video_config


def load_results():
    """加载视频审核结果"""
    results_file = os.path.join(video_config.OUTPUT_DIR, "results.json")
    if not os.path.exists(results_file):
        print(f"结果文件不存在: {results_file}")
        print("请先运行 python video_batch_run.py 完成视频审核")
        return []
    with open(results_file, "r", encoding="utf-8") as f:
        return json.load(f)


def export_csv(results):
    """导出CSV（视频级聚合结果）"""
    output_path = os.path.join(video_config.OUTPUT_DIR, "report.csv")
    headers = [
        "诗名", "文件名", "时长(秒)", "画风", "画风置信度",
        "质量分", "诗词契合度", "史实分",
        "合规", "文字状态", "否决原因",
        "总分", "等级",
        "质量说明", "契合说明", "史实说明",
        "帧数", "Token消耗", "耗时(ms)", "文件大小(MB)",
        "审核状态", "审核时间",
    ]

    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for r in results:
            writer.writerow([
                r.get("poem_name", ""),
                r.get("file_name", ""),
                r.get("duration_sec", 0),
                r.get("style", ""),
                r.get("style_confidence", 0),
                r.get("quality_score", 0),
                r.get("poem_match_score", 0),
                r.get("history_score", 0),
                "合规" if r.get("compliant", True) else "不合规",
                r.get("text_status", ""),
                r.get("veto_reason", ""),
                r.get("total_score", 0),
                r.get("grade", ""),
                r.get("quality_summary", ""),
                r.get("poem_match_summary", ""),
                r.get("history_summary", ""),
                r.get("frame_count", 0),
                r.get("total_tokens", 0),
                r.get("total_latency_ms", 0),
                r.get("file_size_mb", 0),
                r.get("review_status", ""),
                r.get("reviewed_at", ""),
            ])

    print(f"CSV已导出: {output_path}")
    return output_path


def export_excel(results):
    """导出Excel（带格式化，含帧级明细Sheet）"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("缺少openpyxl库，尝试安装...")
        import subprocess
        subprocess.check_call([
            os.path.join(os.path.dirname(sys.executable), "pip"),
            "install", "openpyxl"
        ])
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ============ Sheet 1: 视频审核汇总 ============
    ws = wb.active
    ws.title = "视频审核汇总"

    headers = [
        "诗名", "文件名", "时长(秒)", "画风", "画风置信度",
        "质量分", "诗词契合度", "史实分",
        "合规", "文字状态", "否决原因",
        "总分", "等级",
        "质量说明", "契合说明", "史实说明",
        "帧数", "Token", "耗时(ms)", "大小(MB)",
        "审核状态", "审核时间",
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    grade_fills = {
        "A": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "B": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
        "C": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "D": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }
    grade_fonts = {
        "A": Font(color="006100", bold=True),
        "B": Font(color="0066CC", bold=True),
        "C": Font(color="CC6600", bold=True),
        "D": Font(color="9C0006", bold=True),
    }

    for row_idx, r in enumerate(results, 2):
        grade = r.get("grade", "")
        row_data = [
            r.get("poem_name", ""),
            r.get("file_name", ""),
            r.get("duration_sec", 0),
            r.get("style", ""),
            r.get("style_confidence", 0),
            r.get("quality_score", 0),
            r.get("poem_match_score", 0),
            r.get("history_score", 0),
            "合规" if r.get("compliant", True) else "不合规",
            r.get("text_status", ""),
            r.get("veto_reason", ""),
            r.get("total_score", 0),
            grade,
            r.get("quality_summary", ""),
            r.get("poem_match_summary", ""),
            r.get("history_summary", ""),
            r.get("frame_count", 0),
            r.get("total_tokens", 0),
            r.get("total_latency_ms", 0),
            r.get("file_size_mb", 0),
            r.get("review_status", ""),
            r.get("reviewed_at", ""),
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            # 等级列(13)上色
            if col == 13 and grade in grade_fills:
                cell.fill = grade_fills[grade]
                cell.font = grade_fonts[grade]
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # 总分列(12)上色
            elif col == 12 and grade in grade_fills:
                cell.fill = grade_fills[grade]
            # 史实分列(8)低于7标红
            elif col == 8 and isinstance(val, (int, float)) and val < 7:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.font = Font(color="9C0006", bold=True)

    col_widths = [12, 30, 8, 10, 10, 10, 10, 10, 8, 10, 10, 10, 6, 30, 30, 30, 6, 10, 10, 10, 10, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ============ Sheet 2: 帧级审核明细 ============
    ws2 = wb.create_sheet("帧级明细")

    frame_headers = [
        "视频名", "帧编号", "时间点(秒)", "画风", "画风置信度",
        "质量分", "质量说明", "诗词契合度", "契合说明",
        "史实分", "史实说明", "合规", "文字状态",
        "帧总分", "Token", "耗时(ms)",
    ]

    for col, h in enumerate(frame_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    row = 2
    for r in results:
        if r.get("review_status") != "success":
            continue
        video_name = r.get("file_name", "")
        frame_records = r.get("frame_records", [])
        for fr in frame_records:
            if fr.get("review_status") != "success":
                continue
            text_status = ""
            if fr.get("has_text"):
                tc = fr.get("text_correct", "")
                text_status = tc if tc else "有文字"
            else:
                text_status = "无文字"

            frame_row = [
                video_name,
                fr.get("frame_index", 0),
                fr.get("timestamp", 0),
                fr.get("style", ""),
                fr.get("style_confidence", 0),
                fr.get("quality_score", 0),
                fr.get("quality_detail", ""),
                fr.get("poem_match_score", 0),
                fr.get("poem_match_detail", ""),
                fr.get("history_score", 0),
                fr.get("history_detail", ""),
                "合规" if fr.get("compliant", True) else "不合规",
                text_status,
                fr.get("frame_total_score", 0),
                fr.get("tokens_used", 0),
                fr.get("latency_ms", 0),
            ]
            for col, val in enumerate(frame_row, 1):
                cell = ws2.cell(row=row, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                # 史实分列(10)低于7标红
                if col == 10 and isinstance(val, (int, float)) and val < 7:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(color="9C0006")
            row += 1

    frame_widths = [25, 6, 8, 10, 10, 10, 30, 10, 30, 10, 30, 8, 10, 10, 10, 10]
    for i, w in enumerate(frame_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions

    # ============ Sheet 3: 统计概览 ============
    ws3 = wb.create_sheet("统计概览")

    success = [r for r in results if r.get("review_status") == "success"]
    total = len(success)
    grade_counts = {"A": 0, "B": 0, "C": 0, "D": 0}
    for r in success:
        g = r.get("grade", "C")
        grade_counts[g] = grade_counts.get(g, 0) + 1

    avg_quality = sum(r.get("quality_score", 0) for r in success) / total if total else 0
    avg_poem = sum(r.get("poem_match_score", 0) for r in success) / total if total else 0
    avg_history = sum(r.get("history_score", 0) for r in success) / total if total else 0
    avg_total = sum(r.get("total_score", 0) for r in success) / total if total else 0
    total_tokens = sum(r.get("total_tokens", 0) for r in success)
    non_compliant = sum(1 for r in success if not r.get("compliant", True))
    history_issues = sum(1 for r in success if r.get("history_score", 10) < 7.0)

    stats_data = [
        ["统计项", "数值"],
        ["视频总数", len(results)],
        ["审核成功", total],
        ["审核失败", len(results) - total],
        ["合规数", total - non_compliant],
        ["不合规数", non_compliant],
        ["合规率", f"{(total - non_compliant) / total * 100:.1f}%" if total else "0%"],
        ["史实问题数", history_issues],
        ["", ""],
        ["A级(≥8.0)", grade_counts.get("A", 0)],
        ["B级(≥6.0)", grade_counts.get("B", 0)],
        ["C级(<6.0)", grade_counts.get("C", 0)],
        ["D级(否决)", grade_counts.get("D", 0)],
        ["", ""],
        ["平均质量分", round(avg_quality, 2)],
        ["平均契合度", round(avg_poem, 2)],
        ["平均史实分", round(avg_history, 2)],
        ["平均总分", round(avg_total, 2)],
        ["总Token消耗", total_tokens],
    ]

    for row_idx, (label, val) in enumerate(stats_data, 1):
        ws3.cell(row=row_idx, column=1, value=label).font = Font(bold=(row_idx == 1))
        ws3.cell(row=row_idx, column=2, value=val).font = Font(bold=(row_idx == 1))
        if row_idx == 1:
            for col in [1, 2]:
                ws3.cell(row=row_idx, column=col).fill = header_fill
                ws3.cell(row=row_idx, column=col).font = header_font

    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 15

    # ============ Sheet 4: 异常清单 ============
    ws4 = wb.create_sheet("异常清单")

    ws4.cell(row=1, column=1, value="类型").font = header_font
    ws4.cell(row=1, column=2, value="文件名").font = header_font
    ws4.cell(row=1, column=3, value="说明").font = header_font
    for col in [1, 2, 3]:
        ws4.cell(row=1, column=col).fill = header_fill

    row = 2
    # 审核失败
    for r in [r for r in results if r.get("review_status") == "error"]:
        ws4.cell(row=row, column=1, value="审核失败")
        ws4.cell(row=row, column=2, value=r.get("file_name", ""))
        ws4.cell(row=row, column=3, value=r.get("error_message", ""))
        row += 1

    # 不合规
    for r in [r for r in success if not r.get("compliant", True)]:
        ws4.cell(row=row, column=1, value="不合规")
        ws4.cell(row=row, column=2, value=r.get("file_name", ""))
        ws4.cell(row=row, column=3, value=r.get("veto_detail", "不合规"))
        row += 1

    # 史实问题
    for r in [r for r in success if r.get("history_score", 10) < 7.0]:
        ws4.cell(row=row, column=1, value="史实问题")
        ws4.cell(row=row, column=2, value=r.get("file_name", ""))
        ws4.cell(row=row, column=3, value=r.get("history_summary", ""))
        row += 1

    # 文字乱码
    for r in [r for r in success if r.get("text_status") == "有乱码"]:
        ws4.cell(row=row, column=1, value="文字乱码")
        ws4.cell(row=row, column=2, value=r.get("file_name", ""))
        ws4.cell(row=row, column=3, value="检测到乱码文字")
        row += 1

    ws4.column_dimensions["A"].width = 12
    ws4.column_dimensions["B"].width = 35
    ws4.column_dimensions["C"].width = 50

    # 保存
    output_path = os.path.join(video_config.OUTPUT_DIR, "report.xlsx")
    wb.save(output_path)
    print(f"Excel已导出: {output_path}")
    return output_path


def main():
    """主入口"""
    results = load_results()
    if not results:
        return

    export_csv(results)
    export_excel(results)
    print(f"\n导出完成！文件位于: {video_config.OUTPUT_DIR}")


if __name__ == "__main__":
    main()
